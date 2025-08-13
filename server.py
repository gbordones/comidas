from flask import Flask, request, jsonify
from openpyxl import load_workbook
from datetime import datetime
from flask_cors import CORS # Import CORS

app = Flask(__name__)
CORS(app) # Enable CORS for all routes

@app.route('/')
def index():
    return "Servidor para Tracker de Macros funcionando."

@app.route('/agregar_comida', methods=['POST'])
def agregar_comida():
    data = request.json
    if not data:
        return jsonify({"error": "No se recibieron datos JSON"}), 400

    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tipo_comida = data.get('tipo_comida')
    alimento = data.get('alimento')
    cantidad = data.get('cantidad')
    calorias = data.get('calorias')
    proteinas = data.get('proteinas')
    carbohidratos = data.get('carbohidratos')
    grasas = data.get('grasas')

    try:
        # Cargar el libro de trabajo existente
        wb = load_workbook('alimentos.xlsx')
        ws = wb.active

        # Añadir una nueva fila con los datos
        ws.append([fecha, tipo_comida, alimento, cantidad, calorias, proteinas, carbohidratos, grasas])

        # Guardar el libro de trabajo
        wb.save('alimentos.xlsx')

        return jsonify({"message": "Comida registrada exitosamente"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/obtener_comidas_del_dia', methods=['GET'])
def obtener_comidas_del_dia():
    try:
        wb = load_workbook('alimentos.xlsx')
        ws = wb.active

        today_str = datetime.now().strftime("%Y-%m-%d")
        meals_today = []

        # Iterate through rows, skipping the header (row 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Assuming date is in the first column (index 0)
            # And it's in "YYYY-MM-DD HH:MM:SS" format
            entry_date_str = row[0].split(' ')[0] # Get only the date part

            if entry_date_str == today_str:
                meal = {
                    "fecha": row[0],
                    "tipo_comida": row[1],
                    "alimento": row[2],
                    "cantidad": row[3],
                    "calorias": row[4],
                    "proteinas": row[5],
                    "carbohidratos": row[6],
                    "grasas": row[7]
                }
                meals_today.append(meal)
        
        return jsonify(meals_today), 200
    except FileNotFoundError:
        return jsonify({"error": "Archivo alimentos.xlsx no encontrado."}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5001)
